import os
from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func

app = Flask(__name__)

# ================== SECURITY / CONFIG ==================
app.secret_key = os.environ.get("SECRET_KEY", "local_dev_secret_key_change_me")

app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
if os.environ.get("RENDER") or os.environ.get("FLASK_ENV") == "production":
    app.config["SESSION_COOKIE_SECURE"] = True  # HTTPS üstünden çalışır

# ================== DATABASE (Render Postgres / Local SQLite) ==================
db_url = os.environ.get("DATABASE_URL")
if db_url:
    db_url = db_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = db_url
else:
    os.makedirs("instance", exist_ok=True)
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///instance/verim.db"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# ================= SABİT SEÇENEKLER =================
OPERATIONS = ["1. Operasyon", "2. Operasyon", "Yargı", "Ayar"]

# ================= MODELLER =================
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # admin / operator

class Part(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    part_code = db.Column(db.String(200), nullable=False)   # Parça adı uzun olabilir
    operation = db.Column(db.String(20), nullable=False)    # 4 seçenek
    setup_time = db.Column(db.Integer, nullable=False)      # saniye (parça başına)

    __table_args__ = (
        db.UniqueConstraint("part_code", "operation", name="uq_part_operation"),
    )

class Work(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    operator = db.Column(db.String(50), nullable=False)
    part_code = db.Column(db.String(200), nullable=False)
    operation = db.Column(db.String(20), nullable=False)

    start_time = db.Column(db.String(10), nullable=False)  # HH:MM
    end_time = db.Column(db.String(10), nullable=False)    # HH:MM

    process_time_sec = db.Column(db.Integer, nullable=False)  # saniye (parça başına)
    quantity = db.Column(db.Integer, nullable=False)

    downtime_reason = db.Column(db.String(200))
    downtime_seconds = db.Column(db.Integer, default=0)

    efficiency = db.Column(db.Float, nullable=False)
    date = db.Column(db.DateTime, default=datetime.now)

# ✅ Yeni: Mola tablosu
class BreakTime(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    start_time = db.Column(db.String(5), nullable=False)  # HH:MM
    end_time = db.Column(db.String(5), nullable=False)    # HH:MM

# ================= HESAPLAR =================
def parse_time_hhmm(t: str) -> datetime:
    return datetime.strptime(t, "%H:%M")

def calculate_net_seconds(start_str: str, end_str: str) -> int:
    """
    Net süre (sn) = (end-start) - mola çakışmaları (DB'den okunur)
    """
    start = parse_time_hhmm(start_str)
    end = parse_time_hhmm(end_str)

    total_seconds = int((end - start).total_seconds())
    if total_seconds < 0:
        return 0

    break_seconds = 0

    # ✅ Molaları veritabanından çek
    breaks = BreakTime.query.all()
    for br in breaks:
        bs = parse_time_hhmm(br.start_time)
        be = parse_time_hhmm(br.end_time)

        overlap_start = max(start, bs)
        overlap_end = min(end, be)

        if overlap_start < overlap_end:
            break_seconds += int((overlap_end - overlap_start).total_seconds())

    net = total_seconds - break_seconds
    return max(net, 0)

def calculate_efficiency_seconds(
    start_time: str,
    end_time: str,
    process_sec_per_piece: int,
    setup_sec_per_piece: int,
    quantity: int,
    downtime_seconds: int
) -> float:
    """
    NetSüre = (Bitiş-Başlangıç) - Molalar - KayıpZaman
    StandartParçaSüresi = process + setup (parça başına)
    HedefAdet = NetSüre / StandartParçaSüresi
    Verim% = (GerçekAdet / HedefAdet) * 100
    """
    net_seconds = calculate_net_seconds(start_time, end_time) - max(int(downtime_seconds or 0), 0)

    if net_seconds <= 0:
        return 0.0

    standard_per_piece = int(process_sec_per_piece or 0) + int(setup_sec_per_piece or 0)
    if standard_per_piece <= 0:
        return 0.0

    target_qty = net_seconds / standard_per_piece
    if target_qty <= 0:
        return 0.0

    eff = (int(quantity or 0) / target_qty) * 100
    return round(eff, 2)

def get_part_codes():
    parts = (
        Part.query
        .with_entities(Part.part_code)
        .distinct()
        .order_by(Part.part_code.asc())
        .all()
    )
    return [p.part_code for p in parts]

# ================= ROUTES =================
@app.route("/")
def home():
    return redirect("/login")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session.clear()
            session["user"] = user.username
            session["role"] = user.role
            return redirect("/admin" if user.role == "admin" else "/operator")

        return "Hatalı giriş!"

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/operator", methods=["GET", "POST"])
def operator_panel():
    if session.get("role") != "operator":
        return redirect("/login")

    part_codes = get_part_codes()

    if request.method == "POST":
        part_code = request.form["part_code"].strip()
        operation = request.form["operation"].strip()
        start_time = request.form["start_time"]
        end_time = request.form["end_time"]
        quantity = int(request.form["quantity"])

        process_min = int(request.form["process_min"])
        process_sec_part = int(request.form["process_sec_part"])
        process_sec = process_min * 60 + process_sec_part

        downtime_reason = request.form.get("downtime_reason", "").strip()
        downtime_min = int(request.form.get("downtime_min", "0") or 0)
        downtime_sec_part = int(request.form.get("downtime_sec_part", "0") or 0)
        downtime_seconds = downtime_min * 60 + downtime_sec_part

        if operation not in OPERATIONS:
            return render_template(
                "operator.html",
                operations=OPERATIONS,
                part_codes=part_codes,
                message="Seçenek geçersiz!"
            )

        part = Part.query.filter_by(part_code=part_code, operation=operation).first()
        if not part:
            return render_template(
                "operator.html",
                operations=OPERATIONS,
                part_codes=part_codes,
                message="Bu parça + bu seçenek admin tarafından tanımlı değil!"
            )

        efficiency = calculate_efficiency_seconds(
            start_time=start_time,
            end_time=end_time,
            process_sec_per_piece=process_sec,
            setup_sec_per_piece=part.setup_time,
            quantity=quantity,
            downtime_seconds=downtime_seconds
        )

        new_work = Work(
            operator=session["user"],
            part_code=part_code,
            operation=operation,
            start_time=start_time,
            end_time=end_time,
            process_time_sec=process_sec,
            quantity=quantity,
            downtime_reason=downtime_reason if downtime_reason else None,
            downtime_seconds=downtime_seconds,
            efficiency=efficiency
        )

        db.session.add(new_work)
        db.session.commit()

        return render_template(
            "operator.html",
            operations=OPERATIONS,
            part_codes=part_codes,
            message="Kayıt alınmıştır."
        )

    return render_template("operator.html", operations=OPERATIONS, part_codes=part_codes)

@app.route("/admin", methods=["GET", "POST"])
def admin():
    if session.get("role") != "admin":
        return redirect("/login")

    message = None

    if request.method == "POST":
        action = request.form.get("action", "")

        # Kullanıcı ekle
        if action == "add_user":
            username = request.form["username"].strip()
            password = request.form["password"].strip()
            role = request.form["role"].strip()

            if not username or not password or role not in ("admin", "operator"):
                message = "Kullanıcı bilgileri hatalı!"
            elif User.query.filter_by(username=username).first():
                message = "Bu kullanıcı adı zaten var!"
            else:
                hashed = generate_password_hash(password)
                db.session.add(User(username=username, password_hash=hashed, role=role))
                db.session.commit()
                message = "Kullanıcı eklendi ✅"

        # Parça upsert
        elif action == "upsert_part":
            part_code = request.form["part_code"].strip()
            operation = request.form["operation"].strip()

            setup_min = int(request.form["setup_min"])
            setup_sec_part = int(request.form["setup_sec_part"])
            setup_seconds = setup_min * 60 + setup_sec_part

            if not part_code or operation not in OPERATIONS:
                message = "Parça bilgileri hatalı!"
            else:
                existing = Part.query.filter_by(part_code=part_code, operation=operation).first()
                if existing:
                    existing.setup_time = setup_seconds
                else:
                    db.session.add(Part(part_code=part_code, operation=operation, setup_time=setup_seconds))
                db.session.commit()
                message = "Parça kaydedildi ✅"

        # Excel import
        elif action == "import_excel":
            file = request.files.get("excel_file")
            if not file or file.filename == "":
                message = "Excel dosyası seçilmedi!"
            elif not file.filename.lower().endswith(".xlsx"):
                message = "Sadece .xlsx kabul ediliyor!"
            else:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file)
                    ws = wb.active

                    headers = {}
                    for col in range(1, ws.max_column + 1):
                        val = ws.cell(row=1, column=col).value
                        if val:
                            headers[str(val).strip().upper()] = col

                    required_cols = ["PARÇA ADI", "BİRİNCİ OPERASYON", "İKİNCİ OPERASYON", "YARGI", "AYAR"]
                    for rc in required_cols:
                        if rc not in headers:
                            message = f"Excel başlığı eksik: {rc}"
                            raise ValueError(message)

                    operation_map = {
                        "BİRİNCİ OPERASYON": "1. Operasyon",
                        "İKİNCİ OPERASYON": "2. Operasyon",
                        "YARGI": "Yargı",
                        "AYAR": "Ayar",
                    }

                    added = 0
                    updated = 0
                    skipped = 0

                    for row in range(2, ws.max_row + 1):
                        part_name = ws.cell(row=row, column=headers["PARÇA ADI"]).value
                        if not part_name:
                            continue
                        part_name = str(part_name).strip()

                        for excel_col, system_op in operation_map.items():
                            cell_value = ws.cell(row=row, column=headers[excel_col]).value

                            if cell_value is None or cell_value == "":
                                continue

                            try:
                                setup_seconds = int(float(cell_value))
                            except Exception:
                                skipped += 1
                                continue

                            if setup_seconds <= 0:
                                continue

                            existing = Part.query.filter_by(part_code=part_name, operation=system_op).first()
                            if existing:
                                existing.setup_time = setup_seconds
                                updated += 1
                            else:
                                db.session.add(Part(part_code=part_name, operation=system_op, setup_time=setup_seconds))
                                added += 1

                    db.session.commit()
                    message = f"Excel yüklendi ✅ Eklenen: {added}, Güncellenen: {updated}, Atlanan: {skipped}"

                except Exception as e:
                    if not message:
                        message = f"Excel yükleme hatası: {e}"

        # ✅ Mola ekle
        elif action == "add_break":
            b_start = request.form["break_start"].strip()
            b_end = request.form["break_end"].strip()

            if not b_start or not b_end:
                message = "Mola saatleri boş olamaz!"
            elif b_start >= b_end:
                message = "Mola bitişi başlangıçtan büyük olmalı!"
            else:
                db.session.add(BreakTime(start_time=b_start, end_time=b_end))
                db.session.commit()
                message = "Mola eklendi ✅"

    # Haftalık sıralama (son 7 gün)
    week_start = datetime.now() - timedelta(days=7)
    weekly_rows = (
        db.session.query(
            Work.operator,
            func.avg(Work.efficiency).label("avg_eff"),
            func.count(Work.id).label("cnt")
        )
        .filter(Work.date >= week_start)
        .group_by(Work.operator)
        .order_by(func.avg(Work.efficiency).desc())
        .all()
    )

    weekly_rank = [
        {"operator": r.operator, "avg_eff": round(float(r.avg_eff or 0), 2), "count": int(r.cnt or 0)}
        for r in weekly_rows
    ]

    users = User.query.order_by(User.role.asc(), User.username.asc()).all()
    works = Work.query.order_by(Work.date.desc()).limit(200).all()

    # ✅ Molaları admin ekranına gönder
    breaks = BreakTime.query.order_by(BreakTime.start_time.asc()).all()

    return render_template(
        "admin.html",
        users=users,
        works=works,
        operations=OPERATIONS,
        weekly_rank=weekly_rank,
        breaks=breaks,
        message=message
    )

# ====== ADMIN DELETE ROUTES ======
@app.route("/admin/delete_user/<int:user_id>", methods=["POST"])
def admin_delete_user(user_id):
    if session.get("role") != "admin":
        return redirect("/login")

    user = User.query.get_or_404(user_id)
    if user.username == "admin":
        return "Admin kullanıcı silinemez!"

    db.session.delete(user)
    db.session.commit()
    return redirect("/admin")

@app.route("/admin/delete_work/<int:work_id>", methods=["POST"])
def admin_delete_work(work_id):
    if session.get("role") != "admin":
        return redirect("/login")

    work = Work.query.get_or_404(work_id)
    db.session.delete(work)
    db.session.commit()
    return redirect("/admin")

# ✅ Mola sil
@app.route("/admin/delete_break/<int:break_id>", methods=["POST"])
def admin_delete_break(break_id):
    if session.get("role") != "admin":
        return redirect("/login")

    br = BreakTime.query.get_or_404(break_id)
    db.session.delete(br)
    db.session.commit()
    return redirect("/admin")

# ================= INIT (Render + Gunicorn ile de çalışır) =================
def ensure_db_and_admin():
    with app.app_context():
        db.create_all()

        create_admin = (os.environ.get("ADMIN_CREATE", "false").lower() == "true")
        admin_username = os.environ.get("ADMIN_USERNAME", "admin")
        admin_password = os.environ.get("ADMIN_PASSWORD")

        if create_admin:
            if not admin_password or len(admin_password) < 6:
                return

            if not User.query.filter_by(username=admin_username).first():
                db.session.add(
                    User(
                        username=admin_username,
                        password_hash=generate_password_hash(admin_password),
                        role="admin",
                    )
                )
                db.session.commit()

ensure_db_and_admin()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)